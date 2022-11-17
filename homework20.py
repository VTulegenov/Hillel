import csv
import openpyxl

input_data = []
names = None


with open("task_1.csv", encoding='utf-8') as r_file:
    file_reader = csv.reader(r_file, delimiter=",")
    count = 0
    for row in file_reader:
        if count == 0:
            names = row
        else:
            input_data.append(row)
        count += 1


age_index = [index for index, item in enumerate(names) if item == 'age']
age_index = age_index[0] if age_index else None


if age_index is not None:
    names.pop(age_index)
    for item_list in input_data:
        item_list.pop(age_index)

wb = openpyxl.Workbook()
sheet = wb.active


for item in range(1, len(input_data) + 1):
    cell = sheet.cell(row=1, column=item + 1)
    cell.value = f'person {item}'

for name_index, name in enumerate(names):
    cell = sheet.cell(row=name_index+2, column=1)
    cell.value = name
    for item in range(1, len(input_data) + 1):
        cell = sheet.cell(row=name_index+2, column=item + 1)
        cell.value = input_data[item - 1][name_index]

wb.save('file_xlsx.xlsx')

